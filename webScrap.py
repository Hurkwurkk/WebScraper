from bs4 import BeautifulSoup as bs
import requests
import urllib
import openpyxl
import gui
from tkinter import *
#user data entry

def WebScraper():
    wb = openpyxl.load_workbook('C:\\Users\\hocke\\Desktop\\MyPackage\\data.xlsx')
    ws = wb.active

# guiInput = gui.country.value
    last_row = ws.max_row
    while ws.cell(column=3, row=last_row).value is None and last_row > 0:
        last_row -= 1
    userInput = ws.cell(column=3, row=last_row).value


    print(userInput)
# userInput = user.replace (' ', '+')


    url = f'http://google.com/search?q={userInput}'
    headers_1 = {"user-agent": 'Magic Browser'}

    resp = requests.get(url, headers=headers_1)
    print(resp)

    if resp.status_code == 200:

        soup = bs(resp.content, "html.parser")
        results = []
        resaultVeiw = []

        #print("Before For Loop")
        #print(results)
        #print("The find we are looking for")
        for g in soup.find_all('div'):
            #print("in For Loop")
            anchors = g.find_all('a')
            #print(anchors)
            if anchors:
                link = anchors[0]['href']
                title = g.find('h3')
                item = {
                    "title": title,
                    "link": link,

                }
                veiw = {
                    'https://www.google.com/' + link

                }

                results.append(item)
                resaultVeiw.append(veiw)
                cool = '\n'.join(map(str, resaultVeiw))

        #print("Results")
        #print(*results, sep='\n' + '\n')
        #print(cool)
        #print(*resaultVeiw)




        #---------------------------------
    root = Tk()
    root.configure(background='light blue')
    root.title("WebScraper UI Project")
    root.geometry("850x500")

    heading = Label(root, text="Resaults", bg="light blue").pack()
    instruction = Label(root, text="Copy and paste a URL into your favorite browser", bg="light blue").pack()

    text = Text(root)
    text.insert(INSERT, cool)


    text.pack()
